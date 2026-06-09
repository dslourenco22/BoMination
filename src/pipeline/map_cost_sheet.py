import pandas as pd
import os
import sys
from openpyxl import load_workbook
from pathlib import Path

# Support both script and PyInstaller .exe paths
if getattr(sys, 'frozen', False):
    # Running as PyInstaller executable
    SCRIPT_DIR = Path(sys._MEIPASS)
    OMNI_TEMPLATE_PATH = SCRIPT_DIR / "Files" / "OCTF-1539-COST SHEET.xlsx"
else:
    # Running as script - Files folder is in root directory, go up two levels from src/pipeline/
    SCRIPT_DIR = Path(__file__).parent
    OMNI_TEMPLATE_PATH = SCRIPT_DIR.parent.parent / "Files" / "OCTF-1539-COST SHEET.xlsx"

def generate_output_path(input_file_path, suffix, output_directory=None, extension=".xlsx"):
    """Generate output file path based on input file and optional output directory."""
    input_path = Path(input_file_path)
    base_name = input_path.stem
    output_filename = f"{base_name}{suffix}{extension}"
    
    if output_directory and output_directory.strip():
        output_dir = Path(output_directory)
        output_dir.mkdir(parents=True, exist_ok=True)
        output_path = output_dir / output_filename
    else:
        output_path = input_path.parent / output_filename
    
    return str(output_path)

def get_oem_file():
    while True:
        path = input("Enter path to OEMSecrets Excel file (with prices): ").strip()
        if os.path.isfile(path) and path.endswith(".xlsx"):
            return path
        print("Invalid file path.")

def get_merged_file():
    while True:
        path = input("Enter path to merged Excel file (with DESCRIPTION column): ").strip()
        if os.path.isfile(path) and path.endswith(".xlsx"):
            return path
        print("Invalid merged file path.")

def map_and_insert_data(oem_path, merged_path, template_path=OMNI_TEMPLATE_PATH, output_path=None):
    """Fill the cost sheet template and save it.

    Output location is resolved in this priority order:
      1. `output_path` argument (explicit full path), if given
      2. BOM_COST_SHEET_DIR / BOM_COST_SHEET_NAME environment variables
      3. Default: alongside the source file as "<base>_cost_sheet.xlsx"

    Returns the path the cost sheet was saved to.
    """
    df_oem = pd.read_excel(oem_path, keep_default_na=False, na_values=[''])
    df_merged = pd.read_excel(merged_path, keep_default_na=False, na_values=[''])

    # Get the user's selected company from environment variable
    user_company = os.environ.get("BOM_COMPANY", "").lower()
    print(f"📊 User selected company: {user_company}")

    # Pull description column if present
    if "DESCRIPTION" in df_merged.columns:
        df_oem["DESCRIPTION"] = df_merged["DESCRIPTION"]
    elif "Description" in df_merged.columns:  # NEL format
        df_oem["DESCRIPTION"] = df_merged["Description"]
    else:
        print("⚠️ DESCRIPTION column not found in merged file. Skipping description.")
        df_oem["DESCRIPTION"] = ""

    # Pull PROTON P/N column from merged file for customer part number (NEL only)
    if user_company == "nel":
        if "Proton P/N" in df_merged.columns:
            df_oem["PROTON P/N"] = df_merged["Proton P/N"]
            print("✅ Added PROTON P/N column from merged file (NEL format)")
        else:
            print("⚠️ PROTON P/N column not found in merged file. Skipping customer part number.")
            df_oem["PROTON P/N"] = ""
    else:
        print("📊 Skipping PROTON P/N mapping (not NEL format)")
        df_oem["PROTON P/N"] = ""

    # Pull additional columns from merged file for Primetals format
    if user_company == "primetals":
        # Pull MPN (part number) from merged file
        if "MPN" in df_merged.columns:
            df_oem["MPN"] = df_merged["MPN"]
            print("✅ Added MPN column from merged file (Primetals format)")
        else:
            print("⚠️ MPN column not found in merged file. Skipping part number.")
            df_oem["MPN"] = ""
        
        # Pull QTY (quantity) from merged file
        if "QTY" in df_merged.columns:
            df_oem["QTY"] = df_merged["QTY"]
            print("✅ Added QTY column from merged file (Primetals format)")
        else:
            print("⚠️ QTY column not found in merged file. Skipping quantity.")
            df_oem["QTY"] = ""
        
        # Pull MFG (manufacturer) from merged file
        if "MFG" in df_merged.columns:
            df_oem["MFG"] = df_merged["MFG"]
            print("✅ Added MFG column from merged file (Primetals format)")
        else:
            print("⚠️ MFG column not found in merged file. Skipping manufacturer.")
            df_oem["MFG"] = ""
        
        # Pull ITEM from merged file
        if "ITEM" in df_merged.columns:
            df_oem["ITEM"] = df_merged["ITEM"]
            print("✅ Added ITEM column from merged file (Primetals format)")
        else:
            print("⚠️ ITEM column not found in merged file. Skipping item number.")
            df_oem["ITEM"] = ""

    # Pull additional columns from merged file for Riley Power format
    if user_company == "riley power":
        # Pull MANUFACTURER from merged file
        if "MANUFACTURER" in df_merged.columns:
            df_oem["MANUFACTURER"] = df_merged["MANUFACTURER"]
            print("✅ Added MANUFACTURER column from merged file (Riley Power format)")
        else:
            print("⚠️ MANUFACTURER column not found in merged file. Skipping manufacturer.")
            df_oem["MANUFACTURER"] = ""
        
        # Pull QTY (quantity) from merged file
        if "QTY" in df_merged.columns:
            df_oem["QTY"] = df_merged["QTY"]
            print("✅ Added QTY column from merged file (Riley Power format)")
        else:
            print("⚠️ QTY column not found in merged file. Skipping quantity.")
            df_oem["QTY"] = ""
        
        # Pull MPN (part number) from merged file
        if "MPN" in df_merged.columns:
            df_oem["MPN"] = df_merged["MPN"]
            print("✅ Added MPN column from merged file (Riley Power format)")
        else:
            print("⚠️ MPN column not found in merged file. Skipping part number.")
            df_oem["MPN"] = ""
        
        # Pull ITEM from merged file
        if "ITEM" in df_merged.columns:
            df_oem["ITEM"] = df_merged["ITEM"]
            print("✅ Added ITEM column from merged file (Riley Power format)")
        else:
            print("⚠️ ITEM column not found in merged file. Skipping item number.")
            df_oem["ITEM"] = ""

    # Pull additional columns from merged file for Shanklin format
    if user_company == "shanklin":
        # Pull MPN (part number) from merged file
        if "MPN" in df_merged.columns:
            df_oem["MPN"] = df_merged["MPN"]
            print("✅ Added MPN column from merged file (Shanklin format)")
        else:
            print("⚠️ MPN column not found in merged file. Skipping part number.")
            df_oem["MPN"] = ""
        
        # Pull QTY (quantity) from merged file
        if "QTY" in df_merged.columns:
            df_oem["QTY"] = df_merged["QTY"]
            print("✅ Added QTY column from merged file (Shanklin format)")
        else:
            print("⚠️ QTY column not found in merged file. Skipping quantity.")
            df_oem["QTY"] = ""
        
        # Pull ITEM from merged file
        if "ITEM" in df_merged.columns:
            df_oem["ITEM"] = df_merged["ITEM"]
            print("✅ Added ITEM column from merged file (Shanklin format)")
        else:
            print("⚠️ ITEM column not found in merged file. Skipping item number.")
            df_oem["ITEM"] = ""

    # Pull additional columns from merged file for 901D format
    if user_company == "901d":
        # Pull MPN (part number) from merged file
        if "MPN" in df_merged.columns:
            df_oem["MPN"] = df_merged["MPN"]
            print("✅ Added MPN column from merged file (901D format)")
        else:
            print("⚠️ MPN column not found in merged file. Skipping part number.")
            df_oem["MPN"] = ""
        
        # Pull QTY (quantity) from merged file
        if "QTY" in df_merged.columns:
            df_oem["QTY"] = df_merged["QTY"]
            print("✅ Added QTY column from merged file (901D format)")
        else:
            print("⚠️ QTY column not found in merged file. Skipping quantity.")
            df_oem["QTY"] = ""
        
        # Pull MFR (manufacturer) from merged file
        if "MFR" in df_merged.columns:
            df_oem["MFR"] = df_merged["MFR"]
            print("✅ Added MFR column from merged file (901D format)")
        else:
            print("⚠️ MFR column not found in merged file. Skipping manufacturer.")
            df_oem["MFR"] = ""
        
        # Pull FIND NO. (item number) from merged file
        if "FIND NO." in df_merged.columns:
            df_oem["FIND NO."] = df_merged["FIND NO."]
            print("✅ Added FIND NO. column from merged file (901D format)")
        else:
            print("⚠️ FIND NO. column not found in merged file. Skipping item number.")
            df_oem["FIND NO."] = ""
        
        # Pull 901D P/N (customer part number) from merged file
        if "901D P/N" in df_merged.columns:
            df_oem["901D P/N"] = df_merged["901D P/N"]
            print("✅ Added 901D P/N column from merged file (901D format)")
        else:
            print("⚠️ 901D P/N column not found in merged file. Skipping customer part number.")
            df_oem["901D P/N"] = ""

    # Pull additional columns from merged file for Amazon format
    if user_company == "amazon":
        # Pull Part number from merged file
        if "Part number" in df_merged.columns:
            df_oem["Part number"] = df_merged["Part number"]
            print("✅ Added Part number column from merged file (Amazon format)")
        else:
            print("⚠️ Part number column not found in merged file. Skipping part number.")
            df_oem["Part number"] = ""
        
        # Pull QTY (quantity) from merged file
        if "QTY" in df_merged.columns:
            df_oem["QTY"] = df_merged["QTY"]
            print("✅ Added QTY column from merged file (Amazon format)")
        else:
            print("⚠️ QTY column not found in merged file. Skipping quantity.")
            df_oem["QTY"] = ""
        
        # Pull Manufacturer from merged file
        if "Manufacturer" in df_merged.columns:
            df_oem["Manufacturer"] = df_merged["Manufacturer"]
            print("✅ Added Manufacturer column from merged file (Amazon format)")
        else:
            print("⚠️ Manufacturer column not found in merged file. Skipping manufacturer.")
            df_oem["Manufacturer"] = ""
        
        # Pull Device tag (item number) from merged file
        if "Device tag" in df_merged.columns:
            df_oem["Device tag"] = df_merged["Device tag"]
            print("✅ Added Device tag column from merged file (Amazon format)")
        else:
            print("⚠️ Device tag column not found in merged file. Skipping item number.")
            df_oem["Device tag"] = ""
        
        # Pull Distributor from merged file for supplier notes
        if "Distributor" in df_merged.columns:
            df_oem["Distributor"] = df_merged["Distributor"]
            print("✅ Added Distributor column from merged file (Amazon format)")
        else:
            print("⚠️ Distributor column not found in merged file. Skipping distributor.")
            df_oem["Distributor"] = ""

    # Detect company format based on column names
    oem_columns = df_oem.columns.tolist()
    merged_columns = df_merged.columns.tolist()
    
    print(f"📊 OEM file columns: {oem_columns}")
    print(f"📊 Merged file columns: {merged_columns}")
    
    # Debug: Check for price column variations
    price_columns = [col for col in oem_columns if 'price' in col.lower() or 'cost' in col.lower() or 'usd' in col.lower()]
    print(f"💰 Price-related columns found: {price_columns}")
    
    # Use user's selected company for mapping instead of auto-detection
    if user_company == "nel":
        print("📊 Using NEL format - based on user selection")
        column_mapping = {
            "Internal Reference": "OMNI PART #",
            "Part Number": "COMMERCIAL PART#",
            "Quantity for Single BOM": "UNIT QTY",
            "Extended Quantity for 1 BOM": "EXT QTY",
            "Manufacturer": "MFR",
            "Distributor": "SUPPLIER / NOTES",
            "Minimum Order": "MIN ORDER",
            "Unit Price in USD": "COST EACH",
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "PROTON P/N": "CUST PART #",  # Pull from merged file for NEL
            "DESCRIPTION": "DESCRIPTION"
        }
    elif user_company == "primetals":
        print("📊 Using Primetals format - based on user selection")
        column_mapping = {
            "ITEM": "ITEM",
            "MFG": "MFR",
            "MPN": "COMMERCIAL PART#",
            "QTY": "UNIT QTY",
            "Quantity for Single BOM": "UNIT QTY",  # From OEMsecrets output
            "Unit Price in USD": "COST EACH",
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }
    elif user_company == "riley power":
        print("📊 Using Riley Power format - based on user selection")
        column_mapping = {
            "ITEM": "ITEM",
            "MANUFACTURER": "MFR",  # Use MANUFACTURER from merged file
            "MPN": "COMMERCIAL PART#",  # Use MPN from merged file (converted from Part Number)
            "QTY": "UNIT QTY",  # Use QTY from merged file (converted from Quantity for Single BOM)
            "Quantity for Single BOM": "UNIT QTY",  # From OEMsecrets output
            "Unit Price in USD": "COST EACH",
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }
    elif user_company == "shanklin":
        print("📊 Using Shanklin format - based on user selection")
        column_mapping = {
            "ITEM": "ITEM",
            "MPN": "COMMERCIAL PART#",  # Use MPN from merged file
            "QTY": "UNIT QTY",  # Use QTY from merged file
            "Quantity for Single BOM": "UNIT QTY",  # From OEMsecrets output
            "Unit Price in USD": "COST EACH",
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }
    elif user_company == "901d":
        print("📊 Using 901D format - based on user selection")
        column_mapping = {
            "FIND NO.": "ITEM",  # Use FIND NO. from merged file as item number
            "901D P/N": "CUST PART #",  # Use 901D P/N from merged file as customer part number
            "MFR": "MFR",  # Use MFR from merged file
            "MPN": "COMMERCIAL PART#",  # Use MPN from merged file
            "QTY": "UNIT QTY",  # Use QTY from merged file  
            "Quantity for Single BOM": "UNIT QTY",  # From OEMsecrets output
            "Unit Price in USD": "COST EACH",
            "Unit Price": "COST EACH",  # Alternative price column name
            "Price": "COST EACH",  # Alternative price column name
            "Cost": "COST EACH",  # Alternative price column name
            "Distributor": "SUPPLIER / NOTES",  # Map distributor to supplier field
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }
    elif user_company == "amazon":
        print("📊 Using Amazon format - based on user selection")
        column_mapping = {
            "Device tag": "ITEM",  # Use Device tag from merged file as item number
            "QTY": "UNIT QTY",  # Use QTY from merged file
            "Manufacturer": "MFR",  # Use Manufacturer from merged file
            "Part number": "COMMERCIAL PART#",  # Use Part number from merged file
            "Quantity for Single BOM": "UNIT QTY",  # From OEMsecrets output
            "Unit Price in USD": "COST EACH",
            "Unit Price": "COST EACH",  # Alternative price column name
            "Price": "COST EACH",  # Alternative price column name
            "Cost": "COST EACH",  # Alternative price column name
            "Distributor": "SUPPLIER / NOTES",  # Map distributor to supplier field
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }
    else:
        print("📊 Using Farrell format - based on user selection")

        # Pull key columns from the merged (extracted) BOM file into df_oem.
        # The price-lookup output uses OEMSecrets column names, so MPN and QTY
        # from the original BOM must be grafted in — exactly like other customers.
        mpn_candidates = [
            'MPN', 'Part Number', 'PART NUMBER', 'MODEL NO', 'Model No',
            'CATALOG NUMBER', 'Catalog Number', 'CAT NO', 'Cat No', 'CAT #',
            'ORDER CODE', 'Order Code', 'ORDER NO', 'Order No',
            'ARTICLE NUMBER', 'Article Number', 'ARTICLE NO',
            'PART NO', 'Part No', 'P/N', 'PN',
        ]
        for col in mpn_candidates:
            if col in df_merged.columns:
                df_oem["MPN"] = df_merged[col]
                print(f"✅ Added MPN from merged file column '{col}' (Farrell format)")
                break
        else:
            df_oem["MPN"] = ""
            print("⚠️ MPN column not found in merged file. Skipping part number.")

        qty_candidates = ['QTY', 'Quantity', 'QUANTITY', 'Qty', 'QTY.', 'TOTAL QTY', 'AMOUNT']
        for col in qty_candidates:
            if col in df_merged.columns:
                df_oem["QTY"] = df_merged[col]
                print(f"✅ Added QTY from merged file column '{col}' (Farrell format)")
                break
        else:
            print("⚠️ QTY column not found in merged file. Will use OEMSecrets quantity.")

        mfr_candidates = ['Manufacturer', 'MFR', 'MANUFACTURER', 'MFG', 'BRAND', 'MAKE', 'VENDOR']
        for col in mfr_candidates:
            if col in df_merged.columns:
                df_oem["Manufacturer"] = df_merged[col]
                print(f"✅ Added Manufacturer from merged file column '{col}' (Farrell format)")
                break

        item_candidates = ['ITEM', 'Item', 'ITEM NO', 'Item No', 'LINE', 'LINE NO', 'SEQ', 'SEQ NO']
        for col in item_candidates:
            if col in df_merged.columns:
                df_oem["ITEM"] = df_merged[col]
                print(f"✅ Added ITEM from merged file column '{col}' (Farrell format)")
                break

        cust_pn_candidates = ['Internal Part Number', 'CUST PART', 'CUSTOMER PART', 'INTERNAL P/N',
                               'INTERNAL PART', 'CUSTOMER P/N', 'CUST P/N', 'CUST PN']
        for col in cust_pn_candidates:
            if col in df_merged.columns:
                df_oem["CUST PART #"] = df_merged[col]
                print(f"✅ Added CUST PART # from merged file column '{col}' (Farrell format)")
                break

        if "Description" in df_merged.columns:
            df_oem["DESCRIPTION"] = df_merged["Description"]
        elif "DESCRIPTION" in df_merged.columns:
            df_oem["DESCRIPTION"] = df_merged["DESCRIPTION"]

        # Pull COMMENTS (e.g. "SUPPLIED BY FARREL") → SUPPLIER / NOTES
        comments_candidates = ['COMMENTS', 'Comments', 'COMMENT', 'Comment']
        for col in comments_candidates:
            if col in df_merged.columns:
                df_oem["COMMENTS"] = df_merged[col]
                print(f"✅ Added COMMENTS from merged file column '{col}' (Farrell format)")
                break

        column_mapping = {
            "ITEM": "ITEM",
            "Manufacturer": "MFR",
            "MPN": "COMMERCIAL PART#",
            "CUST PART #": "CUST PART #",
            "QTY": "UNIT QTY",
            "Quantity for Single BOM": "UNIT QTY",
            "Unit Price in USD": "COST EACH",
            "Unit Price": "COST EACH",
            "Price": "COST EACH",
            "Cost": "COST EACH",
            "Minimum Order": "MIN BUY",
            "Lead Time on Additional Stock in Weeks": "LEAD TIME (WEEKS)",
            "Notes": "SUPPLIER / NOTES",
            "COMMENTS": "SUPPLIER / NOTES",
            "DESCRIPTION": "DESCRIPTION"
        }

    print(f"📊 Using column mapping: {column_mapping}")

    # Debug: Show which columns are actually being mapped
    columns_found = {k: v for k, v in column_mapping.items() if k in df_oem.columns}
    columns_missing = {k: v for k, v in column_mapping.items() if k not in df_oem.columns}
    print(f"✅ Columns found for mapping: {columns_found}")
    print(f"❌ Columns missing from OEM file: {columns_missing}")

    # Create a clean mapping that avoids duplicates - take only the first source for each target
    clean_mapping = {}
    for source_col, target_col in column_mapping.items():
        if source_col in df_oem.columns:
            if target_col not in clean_mapping.values():
                clean_mapping[source_col] = target_col
            else:
                print(f"⚠️ Skipping duplicate mapping: {source_col} -> {target_col} (already mapped)")
    
    print(f"📊 Clean mapping (no duplicates): {clean_mapping}")
    
    # Apply the clean mapping
    df_renamed = df_oem.rename(columns=clean_mapping)

    # Get all the target column names we want to keep
    # Always include COST EACH even if price mapping was deduped away
    mapped_target_columns = list(clean_mapping.values())
    if "COST EACH" not in mapped_target_columns and "COST EACH" in df_renamed.columns:
        mapped_target_columns.append("COST EACH")
    df_out = df_renamed[[c for c in mapped_target_columns if c in df_renamed.columns]].copy()
    
    print(f"📊 df_out shape after selection: {df_out.shape}")
    print(f"📊 df_out columns: {list(df_out.columns)}")

    # Handle missing values carefully - preserve numeric columns
    for col in df_out.columns:
        if col == "COST EACH":
            # Convert to numeric; leave NaN as NaN so blank prices write as
            # empty cells rather than 0 (direct injection below may fill them).
            print(f"💰 Processing COST EACH column...")
            print(f"💰 COST EACH sample values: {df_out[col].head().tolist()}")
            df_out[col] = pd.to_numeric(df_out[col], errors='coerce')
            print(f"💰 COST EACH after processing: {df_out[col].tolist()}")
        else:
            # For other columns, replace NaN with "N/A"
            df_out[col] = df_out[col].fillna("N/A")
            df_out[col] = df_out[col].replace("", "N/A")

    # ── TOTAL QTY = UNIT QTY for single-BOM quotes ─────────────────────────────
    if "UNIT QTY" in df_out.columns and "TOTAL QTY" not in df_out.columns:
        df_out["TOTAL QTY"] = df_out["UNIT QTY"]

    # ── Direct price injection ──────────────────────────────────────────────────
    # Bypass the column-mapping pipeline for COST EACH — read the price directly
    # from the prices file so it can never be silently dropped by deduplication.
    if "Unit Price in USD" in df_oem.columns:
        raw_prices = pd.to_numeric(df_oem["Unit Price in USD"], errors='coerce')
        non_zero = (raw_prices > 0).sum()
        print(f"💰 Direct price injection: {non_zero}/{len(raw_prices)} non-zero prices")
        if "COST EACH" not in df_out.columns:
            # NaN stays NaN → blank cell, not 0
            df_out["COST EACH"] = raw_prices.values
        else:
            # Real prices override mapped values; NaN stays NaN → blank cell
            df_out["COST EACH"] = raw_prices.where(raw_prices.notna(), df_out["COST EACH"]).values
    else:
        print("⚠️ Unit Price in USD not found in price file — COST EACH will be blank")

    # Add ITEM numbers (sequential numbering)
    if "ITEM" not in df_out.columns:
        df_out["ITEM"] = range(1, len(df_out) + 1)

    print(f"📊 Final data shape: {df_out.shape}")
    print(f"📊 Final columns: {df_out.columns.tolist()}")
    if len(df_out) > 0:
        print(f"📊 Sample data:\n{df_out.head(2)}")
        # Debug: Show actual COST EACH values
        if "COST EACH" in df_out.columns:
            print(f"💰 COST EACH values: {df_out['COST EACH'].tolist()}")
            print(f"💰 COST EACH data types: {df_out['COST EACH'].dtype}")

    # Validate template file exists before trying to load it
    if not template_path.exists():
        raise FileNotFoundError(
            f"Cost sheet template not found: {template_path}\n\n"
            f"Please ensure the file structure is correct:\n"
            f"BoMination/\n"
            f"  Files/\n"
            f"    OCTF-1539-COST SHEET.xlsx\n"
            f"  src/\n"
            f"    (script files)"
        )

    print(f"LOADING: Loading cost sheet template: {template_path.name}")
    wb = load_workbook(template_path)
    ws = wb.active

    header_row = 12
    excel_headers = {}
    for idx, cell in enumerate(ws[header_row]):
        val = cell.value
        if isinstance(val, str):
            # Clean up headers by normalizing spaces and converting to uppercase
            header = val.strip().upper().replace('\n', ' ').replace('\xa0', ' ')
            # Replace multiple spaces with single spaces
            import re
            header = re.sub(r'\s+', ' ', header)
            excel_headers[header] = idx + 1

    print("Mapped headers from cost sheet:", excel_headers)

    # Clear all data rows in the template before writing so stale content
    # from previous runs doesn't bleed through below our new data
    for clear_row in range(header_row + 1, header_row + 300):
        for col_idx in excel_headers.values():
            ws.cell(row=clear_row, column=col_idx).value = None

    # Insert data row by row
    for i, row in df_out.iterrows():
        for col_name, value in row.items():
            col_upper = col_name.upper().strip()
            if col_upper in excel_headers:
                col_idx = excel_headers[col_upper]
                # Handle different data types appropriately
                if pd.notna(value) and str(value).strip() != '':
                    ws.cell(row=header_row + 1 + i, column=col_idx, value=value)
                    print(f"📝 Inserted '{value}' into {col_upper} at row {header_row + 1 + i}, col {col_idx}")
                else:
                    # Insert empty string for missing/null values
                    ws.cell(row=header_row + 1 + i, column=col_idx, value="")
            else:
                print(f"⚠️ Column '{col_upper}' not found in cost sheet template")

    print(f"✅ Inserted {len(df_out)} rows of data into cost sheet")

    # ── EXT COST formula: = UNIT QTY * COST EACH for every data row (LD)
    from openpyxl.utils import get_column_letter

    _needed = ('UNIT QTY', 'COST EACH', 'EXT COST')
    if all(k in excel_headers for k in _needed):
        uq_letter  = get_column_letter(excel_headers['UNIT QTY'])
        ce_letter  = get_column_letter(excel_headers['COST EACH'])
        ec_col_idx = excel_headers['EXT COST']
        for row_offset in range(len(df_out)):
            data_row = header_row + 1 + row_offset
            ws.cell(row=data_row, column=ec_col_idx,
                    value=f'={uq_letter}{data_row}*{ce_letter}{data_row}')
        print(f"[FORMULA] EXT COST formula written for {len(df_out)} rows")
    else:
        missing = [k for k in _needed if k not in excel_headers]
        print(f"[FORMULA] Skipping EXT COST — columns not found in template: {missing}")

    # ── Resolve where to save the cost sheet ────────────────────────────────────
    oem_path_obj = Path(oem_path)
    base_name = oem_path_obj.stem.replace("_merged_with_prices", "").replace("_merged", "")

    if output_path:
        cost_sheet_path = Path(output_path)
    else:
        out_dir  = os.environ.get("BOM_COST_SHEET_DIR", "").strip()
        out_name = os.environ.get("BOM_COST_SHEET_NAME", "").strip()
        directory = Path(out_dir) if out_dir else oem_path_obj.parent
        filename  = out_name if out_name else f"{base_name}_cost_sheet.xlsx"
        cost_sheet_path = directory / filename

    # Always ensure a .xlsx extension and that the target directory exists
    if cost_sheet_path.suffix.lower() != ".xlsx":
        cost_sheet_path = cost_sheet_path.with_suffix(".xlsx")
    cost_sheet_path.parent.mkdir(parents=True, exist_ok=True)

    print(f"📁 Saving cost sheet to: {cost_sheet_path}")
    wb.save(str(cost_sheet_path))
    print(f"Final cost sheet saved to: {cost_sheet_path}")
    return str(cost_sheet_path)

def main():
    """Main function to be called by the pipeline."""
    oem_path = os.environ.get("OEM_INPUT_PATH")
    merged_path = os.environ.get("MERGED_BOM_PATH")
    
    print(f"🐛 DEBUG: Environment variables:")
    print(f"  OEM_INPUT_PATH: {oem_path}")
    print(f"  MERGED_BOM_PATH: {merged_path}")
    print(f"  OEM file exists: {os.path.exists(oem_path) if oem_path else False}")
    print(f"  Merged file exists: {os.path.exists(merged_path) if merged_path else False}")
    
    if oem_path and os.path.exists(oem_path):
        # Quick check of what's in the OEM file
        try:
            import pandas as pd
            df_test = pd.read_excel(oem_path)
            print(f"🐛 DEBUG: OEM file '{oem_path}' has {df_test.shape[0]} rows and {df_test.shape[1]} columns")
            print(f"🐛 DEBUG: OEM file columns: {list(df_test.columns)}")
        except Exception as e:
            print(f"🐛 DEBUG: Error reading OEM file: {e}")
    
    saved = map_and_insert_data(oem_path, merged_path)
    print(f"✅ Cost sheet written to: {saved}")
    return saved

if __name__ == "__main__":
    main()